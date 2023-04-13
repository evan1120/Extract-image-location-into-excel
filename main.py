import os
import exifread
import openpyxl

dirPath = 'E:/202208湖北武穴老城区旧改/03-过程文件/202203现场调研\武穴现场调研 - 副本'

# 创建一个新的Excel工作表
wb = openpyxl.Workbook()
# 创建一个新的工作表
ws = wb.active

# 添加表头
ws.cell(row=1, column=1, value='img_name')
ws.cell(row=1, column=2, value='Longtitude')
ws.cell(row=1, column=3, value='Latitude')

# 遍历目录下的所有图片文件
for i, img in enumerate(os.listdir(dirPath)):
    img_path = os.path.join(dirPath, img)
    #print(img)

    # 获取GPS数据
    file_path = open(img_path, 'rb')
    contents = exifread.process_file(file_path)
    for key in contents:
        if key == "GPS GPSLongitude":
            longitude = contents["GPS GPSLongitude"].values
        elif key == "GPS GPSLatitude":
            latitude = contents["GPS GPSLatitude"].values

    # 度分秒转换成十进制数据
    longitude_f = longitude[0].num / longitude[0].den + (longitude[1].num / longitude[1].den / 60) + (
                    longitude[2].num / longitude[2].den / 3600)
    latitude_f = latitude[0].num / latitude[0].den + (latitude[1].num / latitude[1].den / 60) + (
                    latitude[2].num / latitude[2].den / 3600)
    #print(longitude_f, latitude_f)

    # 将数据写入工作表中
    ws.cell(row=i+2, column=1, value=i+1)
    ws.cell(row=i+2, column=2, value=img)
    ws.cell(row=i+2, column=3, value=longitude_f)
    ws.cell(row=i+2, column=4, value=latitude_f)

# 保存Excel文件
wb.save('dji_gps_data_Survey.xlsx')
